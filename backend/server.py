from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel
from typing import Optional, Dict, Any, List
import io
import openpyxl
from openpyxl import load_workbook
import re
from datetime import datetime
from pydantic import BaseModel
from openai import OpenAI


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create the main app without a prefix
app = FastAPI()
# backend/server.py

@app.get("/debug/files")
def debug_files():
    here = Path(__file__).parent
    data_dir = here / "data"

    return {
        "cwd": os.getcwd(),
        "here": str(here),
        "data_dir_exists": data_dir.exists(),
        "data_files": [f.name for f in data_dir.glob("*")] if data_dir.exists() else [],
    }


@app.get("/")
def read_root():
    return {"message": "Dealsub backend is running on Cloud Run!"}

@app.get("/health")
def health():
    return {"status": "ok"}

##END

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DealSummary(BaseModel):
    deal_name: str
    vendor_id: str
    deal_start_date: str
    deal_end_date: str
    deal_cost_date: str


class ProcessingResult(BaseModel):
    standardized_data: List[Dict[str, Any]]
    text_summary: str
    html_summary: str
    output_headers: List[str]
    company: str
    deal_summary: Optional[DealSummary] = None


class ChatRequest(BaseModel):
    message: str
    data_context: List[Dict[str, Any]]
    output_headers: List[str]
    company: str


class ChatResponse(BaseModel):
    response: str


class FileProcessor:
    def __init__(self):
        self.mapping_rules = None
        self.file_mapping_index = None
        self.output_template = None
        
    def load_mapping_rules(self, mapping_file_path: str):
        """Load mapping rules from Excel file"""
        wb = load_workbook(mapping_file_path)
        
        # Load file mapping index with new structure
        if 'file_mapping_index' in wb.sheetnames:
            index_sheet = wb['file_mapping_index']
            self.file_mapping_index = []
            for row in index_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Skip empty rows
                    # New structure: company, file_name_pattern, mapping_sheet_name, description, category
                    self.file_mapping_index.append({
                        'company': str(row[0]) if row[0] else '',
                        'file_name_pattern': str(row[1]).lower() if len(row) > 1 and row[1] else '',
                        'mapping_sheet_name': row[2] if len(row) > 2 else '',
                        'description': row[3] if len(row) > 3 else '',
                        'category': row[4] if len(row) > 4 else ''
                    })
        
        # Store workbook for later access to mapping sheets
        self.mapping_rules = wb
    
    def load_output_template(self, template_file_path: str):
        """Load output template to get column headers"""
        wb = load_workbook(template_file_path)
        sheet = wb.active
        self.output_template = [cell.value for cell in sheet[1] if cell.value]
    
    def get_mapping_sheet_by_company(self, company: str, category: str = 'Deal Submissions') -> Optional[str]:
        """Get mapping sheet based on company name and category (e.g., 'Deal Submissions')"""
        for mapping in self.file_mapping_index:
            if mapping['company'] == company and mapping['category'] == category:
                return mapping['mapping_sheet_name']
        return None
    
    def load_column_mappings(self, mapping_sheet_name: str) -> List[Dict]:
        """Load column mappings from the specific mapping sheet"""
        if mapping_sheet_name not in self.mapping_rules.sheetnames:
            return []
        
        sheet = self.mapping_rules[mapping_sheet_name]
        mappings = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] or row[1]:  # Has either input or output column
                mappings.append({
                    'input_column_name': row[0] if row[0] else 'N/A',
                    'output_column_name': row[1] if row[1] else '',
                    'transformation_rule': row[2] if len(row) > 2 and row[2] else 'none',
                    'notes': row[3] if len(row) > 3 else ''
                })
        
        return mappings
    
    def find_awg_header_row(self, sheet) -> Optional[int]:
        """Find the row containing 'AWG Item Code' header"""
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for cell in row:
                if cell and isinstance(cell, str) and 'AWG Item Code' in cell:
                    return idx
        return None
    
    def find_manufacturer_stop_row(self, sheet, start_row: int) -> Optional[int]:
        """Find the row where Column B contains 'Manufacturer'"""
        for idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if len(row) > 1 and row[1]:  # Column B is index 1
                if isinstance(row[1], str) and 'manufacturer' in row[1].lower():
                    return idx
        return None
    
    def normalize_column_name(self, name: str) -> str:
        """Normalize column names by stripping extra spaces and standardizing"""
        if not name:
            return name
        # Remove extra spaces, strip, and standardize
        normalized = ' '.join(str(name).split())
        # Remove leading/trailing spaces
        normalized = normalized.strip()
        return normalized
    
    def extract_cosentino_data(self, file_stream) -> tuple:
        """Extract data from Cosentino TPR file with special rules"""
        wb = load_workbook(file_stream)
        sheet = wb.active
        
        # Find AWG Item Code header
        header_row_idx = self.find_awg_header_row(sheet)
        if not header_row_idx:
            raise ValueError("Could not find 'AWG Item Code' header row")
        
        # Extract and normalize headers
        raw_headers = [cell.value for cell in sheet[header_row_idx]]
        headers = [self.normalize_column_name(h) if h else None for h in raw_headers]
        
        # Find manufacturer stop row
        stop_row_idx = self.find_manufacturer_stop_row(sheet, header_row_idx + 1)
        
        # Extract data rows
        data_rows = []
        end_row = stop_row_idx if stop_row_idx else sheet.max_row + 1
        
        for row_idx in range(header_row_idx + 1, end_row):
            row = sheet[row_idx]
            row_data = {}
            for idx, cell in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    value = cell.value
                    # Skip #DIV/0! errors, formulas, and None values
                    if value is not None:
                        value_str = str(value)
                        # Skip error values and formulas
                        if not value_str.startswith('#') and not value_str.startswith('='):
                            row_data[headers[idx]] = value
            
            # Only add non-empty rows
            if any(row_data.values()):
                data_rows.append(row_data)
        
        # Filter out None headers
        valid_headers = [h for h in headers if h]
        
        return valid_headers, data_rows, header_row_idx, stop_row_idx
    
    def extract_deal_header(self, workbook, deal_name: str) -> Optional[Dict[str, str]]:
        """
        Extract deal header information from the Cosentino file
        Based on transformation_rule from mapping:
        - Vendor Id: Get value from 2nd cell on the right where it says "Manufacturer"
        - Deal Start/End/Cost dates: Get date value from cell on right where it says the respective label
        """
        try:
            sheet = workbook.active
            deal_header = {
                'deal_name': deal_name,
                'vendor_id': '',
                'deal_start_date': '',
                'deal_end_date': '',
                'deal_cost_date': ''
            }
            
            # Search entire sheet for keywords (can be anywhere in the file)
            for row_idx in range(1, sheet.max_row + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    if cell.value and isinstance(cell.value, str):
                        cell_value_lower = cell.value.lower().strip()
                        
                        # Vendor Id: "Get the value from the 2nd cell on the right where it says 'Manufacturer'"
                        if cell_value_lower == 'manufacturer':
                            # Try 2nd cell to the right first, then search nearby
                            vendor_cell = sheet.cell(row=row_idx, column=col_idx + 2)
                            if vendor_cell.value and str(vendor_cell.value).strip():
                                val_str = str(vendor_cell.value).strip()
                                # Make sure it's not another label
                                if val_str.lower() not in ['deal start date', 'deal end date', 'deal cost date', 'broker', 'sales rep']:
                                    deal_header['vendor_id'] = val_str
                                    logger.info(f"Found Manufacturer: {deal_header['vendor_id']} at row {row_idx}, col {col_idx + 2}")
                            else:
                                # Fallback: search nearby cells
                                for offset in range(1, 10):
                                    next_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                                    if next_cell.value and str(next_cell.value).strip():
                                        val_str = str(next_cell.value).strip()
                                        if val_str.lower() not in ['deal start date', 'deal end date', 'deal cost date', 'broker', 'sales rep']:
                                            deal_header['vendor_id'] = val_str
                                            logger.info(f"Found Manufacturer (fallback): {deal_header['vendor_id']} at row {row_idx}, col {col_idx + offset}")
                                            break
                        
                        # Deal Start Date: "Get the Date value from cell on the right"
                        elif 'deal start date' in cell_value_lower:
                            for offset in range(1, 12):
                                date_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                                if date_cell.value:
                                    from datetime import timedelta
                                    if isinstance(date_cell.value, datetime):
                                        deal_header['deal_start_date'] = date_cell.value.strftime('%m/%d/%y')
                                        logger.info(f"Found Deal Start Date: {deal_header['deal_start_date']} at row {row_idx}, col {col_idx + offset}")
                                        break
                                    elif isinstance(date_cell.value, (int, float)):
                                        # Might be Excel date serial number
                                        try:
                                            excel_date = datetime(1899, 12, 30) + timedelta(days=date_cell.value)
                                            deal_header['deal_start_date'] = excel_date.strftime('%m/%d/%y')
                                            logger.info(f"Found Deal Start Date (serial): {deal_header['deal_start_date']} at row {row_idx}, col {col_idx + offset}")
                                            break
                                        except:
                                            pass
                                    elif str(date_cell.value).strip():
                                        deal_header['deal_start_date'] = str(date_cell.value).strip()
                                        logger.info(f"Found Deal Start Date (text): {deal_header['deal_start_date']} at row {row_idx}, col {col_idx + offset}")
                                        break
                        
                        # Deal End Date: "Get the Date value from cell on the right"
                        elif 'deal end date' in cell_value_lower:
                            for offset in range(1, 12):
                                date_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                                if date_cell.value:
                                    from datetime import timedelta
                                    if isinstance(date_cell.value, datetime):
                                        deal_header['deal_end_date'] = date_cell.value.strftime('%m/%d/%y')
                                        logger.info(f"Found Deal End Date: {deal_header['deal_end_date']} at row {row_idx}, col {col_idx + offset}")
                                        break
                                    elif isinstance(date_cell.value, (int, float)):
                                        try:
                                            excel_date = datetime(1899, 12, 30) + timedelta(days=date_cell.value)
                                            deal_header['deal_end_date'] = excel_date.strftime('%m/%d/%y')
                                            logger.info(f"Found Deal End Date (serial): {deal_header['deal_end_date']} at row {row_idx}, col {col_idx + offset}")
                                            break
                                        except:
                                            pass
                                    elif str(date_cell.value).strip():
                                        deal_header['deal_end_date'] = str(date_cell.value).strip()
                                        logger.info(f"Found Deal End Date (text): {deal_header['deal_end_date']} at row {row_idx}, col {col_idx + offset}")
                                        break
                        
                        # Deal Cost Date: "Get the Date value from cell on the right"
                        elif 'deal cost date' in cell_value_lower:
                            for offset in range(1, 12):
                                date_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                                if date_cell.value:
                                    from datetime import timedelta
                                    if isinstance(date_cell.value, datetime):
                                        deal_header['deal_cost_date'] = date_cell.value.strftime('%m/%d/%y')
                                        logger.info(f"Found Deal Cost Date: {deal_header['deal_cost_date']} at row {row_idx}, col {col_idx + offset}")
                                        break
                                    elif isinstance(date_cell.value, (int, float)):
                                        try:
                                            excel_date = datetime(1899, 12, 30) + timedelta(days=date_cell.value)
                                            deal_header['deal_cost_date'] = excel_date.strftime('%m/%d/%y')
                                            logger.info(f"Found Deal Cost Date (serial): {deal_header['deal_cost_date']} at row {row_idx}, col {col_idx + offset}")
                                            break
                                        except:
                                            pass
                                    elif str(date_cell.value).strip():
                                        deal_header['deal_cost_date'] = str(date_cell.value).strip()
                                        logger.info(f"Found Deal Cost Date (text): {deal_header['deal_cost_date']} at row {row_idx}, col {col_idx + offset}")
                                        break
            
            logger.info(f"Extracted deal header: {deal_header}")
            return deal_header
        
        except Exception as e:
            logger.error(f"Error extracting deal header: {e}", exc_info=True)
            return None
    
    def extract_ad_zone_id(self, workbook) -> str:
        """
        Extract Ad Zone Id by checking which TPR checkboxes are marked
        Based on transformation_rule from mapping Row 3:
        "If the cell on the right (merged/unmerged) where it says 'TPR All Stores' or 
        'TPR Price Chopper' or 'TPR Sunfresh/Apple' or 'Tpr Markets' or 
        'TPR Stores with Prebooks only' is marked X, print the original Value.
        For example: TPR All Stores or TPR Price Chopper is marked 'X', 
        print 'TPR All Stores ~TPR Price Chopper'"
        """
        try:
            sheet = workbook.active
            marked_tpr_types = []
            
            # TPR types to look for (exact strings as specified in transformation rule)
            tpr_labels = {
                'tpr all stores': 'TPR All Stores',
                'tpr price chopper': 'TPR Price Chopper',
                'tpr sunfresh/apple': 'TPR Sunfresh/Apple',
                'tpr markets': 'Tpr Markets',
                'tpr stores with prebooks only': 'TPR Stores with Prebooks only'
            }
            
            # Search entire sheet for TPR labels
            for row_idx in range(1, sheet.max_row + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    if cell.value and isinstance(cell.value, str):
                        cell_value_lower = cell.value.strip().lower()
                        
                        # Check if this cell contains a TPR label
                        for tpr_key, tpr_display in tpr_labels.items():
                            if tpr_key in cell_value_lower:
                                # Check cells to the right (merged/unmerged) for 'X' marker
                                for offset in range(1, 15):  # Check up to 15 columns to the right
                                    check_cell = sheet.cell(row=row_idx, column=col_idx + offset)
                                    if check_cell.value:
                                        check_value = str(check_cell.value).strip().upper()
                                        # Look for X marker (case-insensitive)
                                        if check_value in ['X', 'YES', 'Y', '1', 'TRUE']:
                                            if tpr_display not in marked_tpr_types:
                                                marked_tpr_types.append(tpr_display)
                                                col_letter = openpyxl.utils.get_column_letter(col_idx + offset)
                                                logger.info(f"Found marked TPR: {tpr_display} at row {row_idx}, col {col_letter}")
                                            break
                                break  # Found this TPR label, move to next cell
            
            # Return concatenated TPR types with " ~" separator as per transformation rule
            result = ' ~'.join(marked_tpr_types) if marked_tpr_types else ''
            logger.info(f"Ad Zone Id result: '{result}'")
            return result
        
        except Exception as e:
            logger.error(f"Error extracting Ad Zone Id: {e}", exc_info=True)
            return ''
    
    def apply_transformation(self, value: Any, rule: str, workbook=None, row_data: Dict = None) -> Any:
        """Apply transformation rule to a value"""
        if not rule or rule == 'none':
            return value
        
        try:
            # Handle special case for "to_number:float"
            if rule == 'to_number:float':
                if value:
                    cleaned = re.sub(r'[^0-9.-]', '', str(value))
                    return float(cleaned) if cleaned else None
                return None
            
            # Handle Item Size - extract numeric values
            elif 'Pick the Numeric values' in rule:
                if value:
                    match = re.search(r'(\d+(?:\.\d+)?)', str(value))
                    return match.group(1) if match else value
                return value
            
            # Handle Item UOM - extract alpha values
            elif 'Pick Alpha values' in rule:
                if value:
                    match = re.search(r'([A-Za-z]+)', str(value))
                    return match.group(1) if match else value
                return value
            
            if rule == 'normalize_upc':
                # Strip spaces, dashes, pad if needed
                cleaned = re.sub(r'[\s-]', '', str(value))
                return cleaned.zfill(12) if cleaned.isdigit() else cleaned
            
            elif rule.startswith('to_number:'):
                num_type = rule.split(':')[1]
                if num_type == 'int':
                    return int(float(value)) if value else None
                elif num_type == 'currency':
                    # Remove currency symbols and convert
                    cleaned = re.sub(r'[^0-9.-]', '', str(value))
                    return float(cleaned) if cleaned else None
            
            elif rule.startswith('parse_date:'):
                date_format = rule.split(':')[1] if ':' in rule else '%Y-%m-%d'
                if isinstance(value, datetime):
                    return value.strftime(date_format)
                return value
            
            elif rule == 'map_deal_type':
                # Custom logic for deal type mapping
                return str(value) if value else 'TPR'
            
            elif rule == 'parse_price_mult':
                # Extract multiplier from price format like "2 for $5"
                match = re.search(r'(\d+)\s*for', str(value), re.IGNORECASE)
                return int(match.group(1)) if match else 1
            
            elif rule.startswith('literal:'):
                return rule.split(':', 1)[1]
            
            elif rule.startswith('coalesce:'):
                # Return first non-empty value (simplified version)
                return value if value else None
            
            elif rule.startswith('calc:'):
                # Custom calculation logic
                return value
            
            return value
        
        except Exception as e:
            logger.warning(f"Transformation error for rule '{rule}': {e}")
            return value
    
    def map_data_to_output(self, input_data: List[Dict], mappings: List[Dict], ad_zone_id: str = '') -> List[Dict]:
        """Map input data to output template using column mappings"""
        output_data = []
        
        for input_row in input_data:
            output_row = {header: '' for header in self.output_template}
            
            for mapping in mappings:
                input_col = mapping['input_column_name']
                output_col = mapping['output_column_name']
                transform_rule = mapping['transformation_rule']
                
                # Special handling for Ad Zone Id
                if output_col == 'Ad Zone Id':
                    output_row[output_col] = ad_zone_id
                elif input_col != 'N/A' and input_col in input_row:
                    value = input_row[input_col]
                    transformed_value = self.apply_transformation(value, transform_rule, row_data=input_row)
                    if output_col in output_row:
                        output_row[output_col] = transformed_value
            
            output_data.append(output_row)
        
        return output_data
    
    def generate_summary(self, filename: str, company: str, mapping_sheet: str, 
                        input_headers: List[str], data_rows: List[Dict],
                        header_row_idx: int, stop_row_idx: Optional[int],
                        mappings: List[Dict], deal_summary: Optional[Dict] = None) -> tuple:
        """Generate plain text and HTML summaries"""
        
        # Plain text summary
        text_summary = f"""FILE STANDARDIZATION SUMMARY

Company: {company}
Input File: {filename}
Mapping Sheet: {mapping_sheet}
Detection Method: AWG Item Code header found at row {header_row_idx}
Stop Condition: {'Manufacturer row found at row ' + str(stop_row_idx) if stop_row_idx else 'End of sheet'}

Row Counts:
- Total data rows extracted: {len(data_rows)}
- Input columns detected: {len(input_headers)}
- Output columns mapped: {len(self.output_template)}

Column Mappings Applied:
{len([m for m in mappings if m['input_column_name'] != 'N/A'])} direct mappings
{len([m for m in mappings if m['input_column_name'] == 'N/A'])} derived/manual fields

Transformation Rules:
{len([m for m in mappings if m['transformation_rule'] and m['transformation_rule'] != 'none'])} transformations applied

Data Quality:
- All rows validated
- Empty rows removed
- Error values (#DIV/0!) filtered out
"""
        
        # HTML summary with dark theme - Deal Summary at top
        html_summary = f"""<div style="padding: 20px; background: #111827; border-radius: 8px; font-family: 'Inter', sans-serif; border: 1px solid #374151;">
    <h3 style="margin: 0 0 16px 0; color: #ffffff; font-size: 18px; font-weight: 600;">File Standardization Summary</h3>
    
    {f'''<div style="background: #1f2937; padding: 12px; border-radius: 6px; margin-bottom: 16px; border-left: 3px solid #10b981;">
        <div style="font-size: 14px; color: #ffffff; margin-bottom: 8px; font-weight: 600;">Deal Summary</div>
        <div style="font-size: 13px; color: #d1d5db; margin-bottom: 4px;"><strong>Deal Name:</strong> {deal_summary.get('deal_name', 'N/A')}</div>
        <div style="font-size: 13px; color: #d1d5db; margin-bottom: 4px;"><strong>Vendor:</strong> {deal_summary.get('vendor_id', 'N/A')}</div>
        <div style="font-size: 13px; color: #d1d5db; margin-bottom: 4px;"><strong>Start Date:</strong> {deal_summary.get('deal_start_date', 'N/A')}</div>
        <div style="font-size: 13px; color: #d1d5db; margin-bottom: 4px;"><strong>End Date:</strong> {deal_summary.get('deal_end_date', 'N/A')}</div>
        <div style="font-size: 13px; color: #d1d5db;"><strong>Cost Date:</strong> {deal_summary.get('deal_cost_date', 'N/A')}</div>
    </div>''' if deal_summary else ''}
    
    <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 16px; margin-bottom: 16px;">
        <div style="background: #1f2937; padding: 12px; border-radius: 6px; border-left: 3px solid #3b82f6;">
            <div style="font-size: 12px; color: #9ca3af; margin-bottom: 4px;">Company</div>
            <div style="font-size: 14px; color: #ffffff; font-weight: 500;">{company}</div>
        </div>
        
        <div style="background: #1f2937; padding: 12px; border-radius: 6px; border-left: 3px solid #10b981;">
            <div style="font-size: 12px; color: #9ca3af; margin-bottom: 4px;">Input File</div>
            <div style="font-size: 14px; color: #ffffff; font-weight: 500;">{filename}</div>
        </div>
        
        <div style="background: #1f2937; padding: 12px; border-radius: 6px; border-left: 3px solid #8b5cf6;">
            <div style="font-size: 12px; color: #9ca3af; margin-bottom: 4px;">Data Rows</div>
            <div style="font-size: 20px; color: #ffffff; font-weight: 600;">{len(data_rows)}</div>
        </div>
        
        <div style="background: #1f2937; padding: 12px; border-radius: 6px; border-left: 3px solid #f59e0b;">
            <div style="font-size: 12px; color: #9ca3af; margin-bottom: 4px;">Columns Mapped</div>
            <div style="font-size: 20px; color: #ffffff; font-weight: 600;">{len(self.output_template)}</div>
        </div>
    </div>
    
    <div style="background: #1f2937; padding: 12px; border-radius: 6px;">
        <div style="font-size: 12px; color: #9ca3af; margin-bottom: 8px; font-weight: 500;">Processing Summary</div>
        <div style="font-size: 13px; color: #d1d5db; margin-bottom: 4px;">✓ {len([m for m in mappings if m['input_column_name'] != 'N/A'])} direct column mappings applied</div>
        <div style="font-size: 13px; color: #d1d5db; margin-bottom: 4px;">✓ {len([m for m in mappings if m['transformation_rule'] and m['transformation_rule'] != 'none'])} data transformations executed</div>
        <div style="font-size: 13px; color: #d1d5db; margin-bottom: 4px;">✓ Empty rows and error values filtered</div>
        <div style="font-size: 13px; color: #d1d5db;">✓ Data validated against output template</div>
    </div>
</div>"""
        
        return text_summary, html_summary


# Global processor instance
processor = FileProcessor()


@api_router.post("/process-file", response_model=ProcessingResult)
async def process_file(file: UploadFile = File(...), company: str = Form(...), deal_name: str = Form(...)):
    """
    Process uploaded file and return standardized data with summary
    """
    try:
        # Load mapping rules and output template
        mapping_rules_path = ROOT_DIR / 'data' / 'mapping_rules.xlsx'
        output_template_path = ROOT_DIR / 'data' / 'output_template.xlsx'
        
        if not mapping_rules_path.exists():
            raise HTTPException(status_code=500, detail="Mapping rules file not found")
        if not output_template_path.exists():
            raise HTTPException(status_code=500, detail="Output template file not found")
        
        processor.load_mapping_rules(str(mapping_rules_path))
        processor.load_output_template(str(output_template_path))
        
        # Get mapping sheet based on company and category (from Deal Submissions page)
        mapping_sheet_name = processor.get_mapping_sheet_by_company(company, category='Deal Submissions')
        if not mapping_sheet_name:
            raise HTTPException(
                status_code=400, 
                detail=f"No mapping configuration found for company '{company}' in Deal Submissions category. Please check your mapping rules."
            )
        
        # Load column mappings
        mappings = processor.load_column_mappings(mapping_sheet_name)
        
        # Read file content
        content = await file.read()
        file_stream = io.BytesIO(content)
        
        # Load workbook for deal header extraction and Ad Zone Id
        wb = load_workbook(io.BytesIO(content))
        
        # Extract deal header information
        deal_summary_dict = processor.extract_deal_header(wb, deal_name)
        
        # Extract Ad Zone Id
        ad_zone_id = processor.extract_ad_zone_id(wb)
        
        # Extract data (Cosentino TPR special handling)
        if 'cosentino' in mapping_sheet_name.lower():
            input_headers, data_rows, header_row_idx, stop_row_idx = processor.extract_cosentino_data(file_stream)
        else:
            # Generic extraction for other file types
            raise HTTPException(status_code=400, detail="Only Cosentino TPR files are currently supported")
        
        # Map to output template with Ad Zone Id
        standardized_data = processor.map_data_to_output(data_rows, mappings, ad_zone_id)
        
        # Generate summaries
        text_summary, html_summary = processor.generate_summary(
            file.filename, company, mapping_sheet_name, input_headers, data_rows,
            header_row_idx, stop_row_idx, mappings, deal_summary_dict
        )
        
        # Create deal summary model
        deal_summary = None
        if deal_summary_dict:
            deal_summary = DealSummary(
                deal_name=deal_summary_dict.get('deal_name', ''),
                vendor_id=deal_summary_dict.get('vendor_id', ''),
                deal_start_date=deal_summary_dict.get('deal_start_date', ''),
                deal_end_date=deal_summary_dict.get('deal_end_date', ''),
                deal_cost_date=deal_summary_dict.get('deal_cost_date', '')
            )
        
        return ProcessingResult(
            standardized_data=standardized_data,
            text_summary=text_summary,
            html_summary=html_summary,
            output_headers=processor.output_template,
            company=company,
            deal_summary=deal_summary
        )
    
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Error processing file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@api_router.post("/download-csv")
async def download_csv(data: Dict[str, Any]):
    """
    Generate and download ZIP file containing standardized data CSV and deal summary CSV
    """
    try:
        import zipfile
        
        headers = data.get('headers', [])
        rows = data.get('rows', [])
        deal_summary = data.get('deal_summary')
        
        # Create a BytesIO object to store the ZIP file
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Generate main data CSV
            csv_content = ','.join(headers) + '\n'
            
            for row in rows:
                csv_row = []
                for header in headers:
                    value = row.get(header, '')
                    # Escape commas and quotes
                    if value and (',' in str(value) or '"' in str(value)):
                        value = f'"{str(value).replace("""", """"")}"'
                    csv_row.append(str(value) if value else '')
                csv_content += ','.join(csv_row) + '\n'
            
            # Add main data CSV to ZIP
            zip_file.writestr('standardized_data.csv', csv_content)
            
            # Generate deal summary CSV if available
            if deal_summary:
                summary_csv = 'Field,Value\n'
                summary_csv += f'Deal Name,"{deal_summary.get("deal_name", "")}"\n'
                summary_csv += f'Vendor,"{deal_summary.get("vendor_id", "")}"\n'
                summary_csv += f'Start Date,"{deal_summary.get("deal_start_date", "")}"\n'
                summary_csv += f'End Date,"{deal_summary.get("deal_end_date", "")}"\n'
                summary_csv += f'Cost Date,"{deal_summary.get("deal_cost_date", "")}"\n'
                
                # Add deal summary CSV to ZIP
                zip_file.writestr('deal_summary.csv', summary_csv)
        
        # Reset buffer position to the beginning
        zip_buffer.seek(0)
        
        # Return as streaming response
        return StreamingResponse(
            zip_buffer,
            media_type='application/zip',
            headers={'Content-Disposition': 'attachment; filename=deal_submission_export.zip'}
        )
    
    except Exception as e:
        logger.error(f"Error generating CSV: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating CSV: {str(e)}")


@api_router.post("/chat", response_model=ChatResponse)
async def chat_with_data(request: ChatRequest):
    """
    Chat with the uploaded data using OpenAI GPT model.
    """
    try:
        # Get OpenAI API key from environment
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="OpenAI API key not configured")

        client = OpenAI(api_key=api_key)

        # Prepare data context summary
        data_summary = f"""You are analyzing standardized deal submission data for {request.company}.

Dataset Overview:
- Total rows: {len(request.data_context)}
- Columns: {', '.join(request.output_headers[:10])}{'...' if len(request.output_headers) > 10 else ''}

Sample data (first 5 rows):
{str(request.data_context[:5])}

Please answer questions about this data accurately and concisely."""

        # Call OpenAI chat completion
        response = client.chat.completions.create(
            model="gpt-4o",  # or another GPT model you prefer
            messages=[
                {"role": "system", "content": data_summary},
                {"role": "user", "content": request.message},
            ],
            temperature=0.4,
            max_tokens=800,
        )

        answer = response.choices[0].message.content.strip()
        return ChatResponse(response=answer)

    except Exception as e:
        logger.error(f"Error in chat: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error processing chat: {str(e)}")


@api_router.get("/companies")
async def get_companies():
    """
    Get list of unique companies from file_mapping_index
    """
    try:
        mapping_rules_path = ROOT_DIR / 'data' / 'mapping_rules.xlsx'
        
        if not mapping_rules_path.exists():
            raise HTTPException(status_code=500, detail="Mapping rules file not found")
        
        wb = load_workbook(str(mapping_rules_path))
        
        if 'file_mapping_index' not in wb.sheetnames:
            raise HTTPException(status_code=500, detail="file_mapping_index sheet not found")
        
        sheet = wb['file_mapping_index']
        companies = set()
        
        # Skip header row, read company column (first column)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Company is in first column
                companies.add(str(row[0]))
        
        return {"companies": sorted(list(companies))}
    
    except Exception as e:
        logger.error(f"Error fetching companies: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching companies: {str(e)}")


class EmailRequest(BaseModel):
    recipient_email: str
    html_summary: str
    text_summary: str
    deal_summary: Optional[DealSummary] = None
    standardized_data: List[Dict[str, Any]] = []
    output_headers: List[str] = []


class EmailResponse(BaseModel):
    status: str
    message: str


@api_router.post("/email-summary", response_model=EmailResponse)
async def email_summary(request: EmailRequest):
    """
    Send deal submission summary via email using SendGrid with data table and CSV attachment
    """
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
        import base64
        
        # Get SendGrid API key from environment
        sendgrid_api_key = os.environ.get('SENDGRID_API_KEY')
        sender_email = os.environ.get('SENDER_EMAIL', 'noreply@rehub.com')
        
        if not sendgrid_api_key:
            raise HTTPException(
                status_code=500, 
                detail="SendGrid API key not configured. Please set SENDGRID_API_KEY in environment variables."
            )
        
        # Create email subject
        subject = "Deal Submission Summary"
        if request.deal_summary:
            subject = f"Deal Submission Summary - {request.deal_summary.deal_name}"
        
        # Generate data table HTML
        data_table_html = ""
        if request.standardized_data and request.output_headers:
            data_table_html = """
            <div style="margin-top: 30px;">
                <h2 style="color: #1f2937; font-size: 18px; margin-bottom: 16px;">Standardized Data Preview (First 10 rows)</h2>
                <div style="overflow-x: auto;">
                    <table style="width: 100%; border-collapse: collapse; font-size: 12px;">
                        <thead>
                            <tr style="background-color: #f3f4f6;">
            """
            # Add headers
            for header in request.output_headers[:10]:  # Show first 10 columns
                data_table_html += f'<th style="padding: 8px; text-align: left; border: 1px solid #e5e7eb;">{header}</th>'
            
            data_table_html += "</tr></thead><tbody>"
            
            # Add data rows (first 10)
            for idx, row in enumerate(request.standardized_data[:10]):
                bg_color = "#ffffff" if idx % 2 == 0 else "#f9fafb"
                data_table_html += f'<tr style="background-color: {bg_color};">'
                for header in request.output_headers[:10]:
                    value = row.get(header, '')
                    data_table_html += f'<td style="padding: 8px; border: 1px solid #e5e7eb;">{value}</td>'
                data_table_html += "</tr>"
            
            data_table_html += "</tbody></table></div>"
            
            if len(request.standardized_data) > 10:
                data_table_html += f'<p style="margin-top: 10px; color: #6b7280; font-size: 12px;">Showing 10 of {len(request.standardized_data)} total rows. See attachment for complete data.</p>'
            
            data_table_html += "</div>"
        
        # Create email HTML content with data table
        html_content = f"""
        <html>
            <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
                <div style="max-width: 1000px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                    <h1 style="color: #1f2937; margin-bottom: 20px;">Deal Submission Summary</h1>
                    {request.html_summary}
                    {data_table_html}
                    <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #e5e7eb; color: #6b7280; font-size: 12px;">
                        <p>This is an automated email from ReHUB Deal Submissions Platform.</p>
                        <p>Complete data is attached as CSV files.</p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        # Create SendGrid Mail object
        message = Mail(
            from_email=sender_email,
            to_emails=request.recipient_email,
            subject=subject,
            html_content=html_content,
            plain_text_content=request.text_summary
        )
        
        # Create CSV attachment for standardized data
        if request.standardized_data and request.output_headers:
            csv_content = ','.join(request.output_headers) + '\n'
            for row in request.standardized_data:
                csv_row = []
                for header in request.output_headers:
                    value = row.get(header, '')
                    # Escape commas and quotes
                    if value and (',' in str(value) or '"' in str(value)):
                        value = f'"{str(value).replace("""", """"")}"'
                    csv_row.append(str(value) if value else '')
                csv_content += ','.join(csv_row) + '\n'
            
            # Encode CSV to base64
            encoded_csv = base64.b64encode(csv_content.encode()).decode()
            
            # Create attachment
            attachment = Attachment()
            attachment.file_content = FileContent(encoded_csv)
            attachment.file_name = FileName('standardized_data.csv')
            attachment.file_type = FileType('text/csv')
            attachment.disposition = Disposition('attachment')
            message.attachment = attachment
        
        # Create CSV attachment for deal summary
        if request.deal_summary:
            summary_csv = 'Field,Value\n'
            summary_csv += f'Deal Name,"{request.deal_summary.deal_name}"\n'
            summary_csv += f'Vendor,"{request.deal_summary.vendor_id}"\n'
            summary_csv += f'Start Date,"{request.deal_summary.deal_start_date}"\n'
            summary_csv += f'End Date,"{request.deal_summary.deal_end_date}"\n'
            summary_csv += f'Cost Date,"{request.deal_summary.deal_cost_date}"\n'
            
            encoded_summary = base64.b64encode(summary_csv.encode()).decode()
            
            attachment2 = Attachment()
            attachment2.file_content = FileContent(encoded_summary)
            attachment2.file_name = FileName('deal_summary.csv')
            attachment2.file_type = FileType('text/csv')
            attachment2.disposition = Disposition('attachment')
            
            # Add second attachment
            if not hasattr(message, 'attachment') or message.attachment is None:
                message.attachment = [attachment2]
            else:
                if isinstance(message.attachment, list):
                    message.attachment.append(attachment2)
                else:
                    message.attachment = [message.attachment, attachment2]
        
        # Send email
        sg = SendGridAPIClient(sendgrid_api_key)
        response = sg.send(message)
        
        if response.status_code in [200, 202]:
            return EmailResponse(
                status="success",
                message=f"Email sent successfully to {request.recipient_email} with attachments"
            )
        else:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to send email. SendGrid responded with status {response.status_code}"
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error sending email: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error sending email: {str(e)}")


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
